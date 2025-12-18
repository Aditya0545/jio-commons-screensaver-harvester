#!/usr/bin/env python3
"""
Wikimedia Commons Quality Image Harvester - Set-Top Box Version
Focused on: Wiki Loves Monuments 2025, Wiki Loves Folklore, Wiki Loves Birds
Optimized with: ±20px tolerance, dual-file output, API optimizations
"""

import sys
import os
import argparse
import re
import shutil
import time
from typing import Optional, List, Dict, Tuple, Set
from collections import deque
from urllib.parse import urlparse, unquote, parse_qs, quote
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import config

# ========== GLOBAL SESSION FOR CONNECTION POOLING ==========
SESSION = requests.Session()
SESSION.headers.update({
    "Accept": "application/json",
    "Accept-Encoding": "gzip, deflate",
})

# ========== LOAD ALL SETTINGS FROM CONFIG ONLY ==========
try:
    USER_AGENT = f"Wiki-Jio/1.0 (MediaWiki user: {config.MEDIAWIKI_USERNAME})"
    SESSION.headers["User-Agent"] = USER_AGENT
except Exception:
    print("[X] config.py missing or invalid")
    sys.exit(1)

if not getattr(config, "MEDIAWIKI_USERNAME", "").strip():
    print("[X] MEDIAWIKI_USERNAME is empty in config.py")
    sys.exit(1)


def api_request(base_url: str, params: dict, timeout: int = None, retries: int = 3) -> Optional[dict]:
    """Make API request with retries, gzip compression, and connection pooling."""
    if timeout is None:
        timeout = getattr(config, "API_TIMEOUT", 30)
    
    api_url = f"{base_url}/w/api.php"
    
    # Add maxlag parameter to respect server load
    params["maxlag"] = 5
    
    for attempt in range(retries):
        try:
            resp = SESSION.get(api_url, params=params, timeout=timeout)
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.Timeout:
            if attempt < retries - 1:
                time.sleep(1 * (attempt + 1))  # Exponential backoff
                continue
            print(f"    [!] API timeout after {retries} attempts")
            return None
        except requests.exceptions.RequestException as e:
            if attempt < retries - 1:
                time.sleep(1 * (attempt + 1))
                continue
            print(f"    [!] API error: {e}")
            return None
        except Exception:
            if attempt < retries - 1:
                continue
            return None


def extract_category_title(url: str) -> Optional[Tuple[str, str]]:
    """Extract base_url and category title from URL."""
    parsed = urlparse(url)
    base_url = f"{parsed.scheme}://{parsed.netloc}"
    
    if "/wiki/" in parsed.path:
        cat_title = unquote(parsed.path.split("/wiki/", 1)[1])
    elif "title=" in parsed.query:
        query_params = parse_qs(parsed.query)
        cat_title = unquote(query_params.get("title", [""])[0])
    else:
        return None
    
    if not cat_title.startswith("Category:"):
        return None
    
    return base_url, cat_title


def is_winning_category(cat_title: str) -> bool:
    """Check if category contains winning/award images."""
    cat_lower = cat_title.lower()
    for keyword in getattr(config, "WINNING_KEYWORDS", []):
        if keyword.lower() in cat_lower:
            return True
    return False


def should_queue_category(cat_title: str) -> bool:
    cat_lower = cat_title.lower()

    exclude = getattr(config, "SUBCATEGORY_EXCLUDE_KEYWORDS", []) or []
    for kw in exclude:
        if kw and kw.lower() in cat_lower:
            return False

    include = getattr(config, "SUBCATEGORY_INCLUDE_KEYWORDS", []) or []
    if include:
        for kw in include:
            if kw and kw.lower() in cat_lower:
                return True
        return False

    return True


def category_priority_key(cat_title: str) -> int:
    cat_lower = cat_title.lower()
    if is_winning_category(cat_title):
        return 0
    for kw in getattr(config, "SUBCATEGORY_HIGH_PRIORITY_KEYWORDS", []) or []:
        if kw and kw.lower() in cat_lower:
            return 1
    for kw in getattr(config, "SUBCATEGORY_INCLUDE_KEYWORDS", []) or []:
        if kw and kw.lower() in cat_lower:
            return 2
    return 3


def resolution_score(width: int, height: int, targets: List[Tuple[int, int]], tolerance: float) -> float:
    """Score how well image resolution matches targets. Returns 0-100."""
    best_score = 0
    
    for tw, th in targets:
        w_diff = abs(width - tw) / tw
        h_diff = abs(height - th) / th
        
        if w_diff <= tolerance and h_diff <= tolerance:
            score = 100 - (w_diff + h_diff) * 50
            best_score = max(best_score, score)
        
        target_aspect = tw / th
        actual_aspect = width / height
        aspect_diff = abs(target_aspect - actual_aspect) / target_aspect
        
        if aspect_diff <= 0.0025:
            if width >= tw and height >= th:
                best_score = max(best_score, 95)
            else:
                size_ratio = (width * height) / (tw * th)
                if size_ratio >= 0.5:
                    score = 80 + min(10, max(0, (size_ratio - 0.5) * 20))
                    best_score = max(best_score, score)
    
    return best_score


def resolution_score_pixels(width: int, height: int, targets: List[Tuple[int, int]], pixel_tolerance: int) -> float:
    """Score resolution match with pixel tolerance (±N pixels). Returns 0-100."""
    best_score = 0
    if pixel_tolerance < 0:
        pixel_tolerance = 0
    
    for tw, th in targets:
        dw = abs(width - tw)
        dh = abs(height - th)
        
        if pixel_tolerance == 0:
            if dw == 0 and dh == 0:
                return 100
            continue
        
        if dw <= pixel_tolerance and dh <= pixel_tolerance:
            closeness = (dw + dh) / (2 * pixel_tolerance)
            score = 100 - closeness * 5
            best_score = max(best_score, score)
    
    return best_score


def clean_html(text: str) -> str:
    """Remove HTML tags from text."""
    if not text:
        return ""
    text = re.sub(r'<[^>]+>', '', text)
    text = ' '.join(text.split())
    return text.strip()


def fetch_batch_metadata(base_url: str, titles: List[str]) -> Dict[str, dict]:
    """Fetch metadata for up to 50 images in one API call."""
    if not titles:
        return {}
    
    titles_str = "|".join(titles[:50])
    
    params = {
        "action": "query",
        "titles": titles_str,
        "prop": "imageinfo",
        "iiprop": "user|size|url|extmetadata",
        "format": "json",
        "formatversion": "2",
    }
    
    data = api_request(base_url, params)
    if not data:
        return {}
    
    pages = data.get("query", {}).get("pages", [])
    result = {}
    
    for page in pages:
        if page.get("missing"):
            continue
        
        imageinfo = page.get("imageinfo", [])
        if not imageinfo:
            continue
        
        info = imageinfo[0]
        title = page.get("title", "")
        
        ext = info.get("extmetadata", {}) or {}
        
        def ext_value(key: str) -> Optional[str]:
            val = ext.get(key)
            if isinstance(val, dict):
                return val.get("value")
            return val
        
        width = info.get("width")
        height = info.get("height")
        
        if not width or not height:
            continue
        
        if (
            width < getattr(config, "MIN_WIDTH", 800)
            or height < getattr(config, "MIN_HEIGHT", 600)
            or width > getattr(config, "MAX_WIDTH", 10000)
            or height > getattr(config, "MAX_HEIGHT", 10000)
        ):
            continue
        
        wiki_title = title.replace(" ", "_")
        page_url = f"{base_url}/wiki/{quote(wiki_title, safe=':_()%')}"
        file_url = info.get("url")
        
        author = clean_html(ext_value("Artist") or info.get("user") or "Unknown")
        license_type = clean_html(ext_value("LicenseShortName") or ext_value("License") or "Unknown")
        description = clean_html(ext_value("ImageDescription") or ext_value("ObjectName") or "")
        
        result[title] = {
            "page_url": page_url,
            "title": title,
            "url": file_url,
            "author": author,
            "size_bytes": info.get("size"),
            "resolution": f"{width}x{height}",
            "license_type": license_type,
            "description": description,
            "width": width,
            "height": height,
        }
    
    return result


def fetch_batch_sizes(base_url: str, titles: List[str]) -> Dict[str, dict]:
    """Fetch width/height (and bytes) for up to 50 images in one API call."""
    if not titles:
        return {}
    
    titles_str = "|".join(titles[:50])
    
    params = {
        "action": "query",
        "titles": titles_str,
        "prop": "imageinfo",
        "iiprop": "size",
        "format": "json",
        "formatversion": "2",
    }
    
    data = api_request(base_url, params)
    if not data:
        return {}
    
    pages = data.get("query", {}).get("pages", [])
    result = {}
    
    for page in pages:
        if page.get("missing"):
            continue
        
        imageinfo = page.get("imageinfo", [])
        if not imageinfo:
            continue
        
        info = imageinfo[0]
        title = page.get("title", "")
        width = info.get("width")
        height = info.get("height")
        
        if not title or not width or not height:
            continue
        
        if (
            width < getattr(config, "MIN_WIDTH", 800)
            or height < getattr(config, "MIN_HEIGHT", 600)
            or width > getattr(config, "MAX_WIDTH", 10000)
            or height > getattr(config, "MAX_HEIGHT", 10000)
        ):
            continue
        
        result[title] = {
            "title": title,
            "width": width,
            "height": height,
            "size_bytes": info.get("size"),
        }
    
    return result


def get_thumbnail_url(base_url: str, title: str) -> Optional[dict]:
    """Get thumbnail URL for a specific width."""
    params = {
        "action": "query",
        "titles": title,
        "prop": "imageinfo",
        "iiprop": "url|size",
        "iiurlwidth": str(getattr(config, "THUMBNAIL_WIDTH", 1920)),
        "format": "json",
        "formatversion": "2",
    }
    
    data = api_request(base_url, params, timeout=10)
    if not data:
        return None
    
    pages = data.get("query", {}).get("pages", [])
    if not pages:
        return None
    
    page = pages[0]
    imageinfo = page.get("imageinfo", [])
    if not imageinfo:
        return None
    
    info = imageinfo[0]
    thumb_url = info.get("thumburl") or info.get("url")
    
    return {
        "url": thumb_url,
        "width": info.get("thumbwidth"),
        "height": info.get("thumbheight"),
    }


def process_thumbnails_batch(base_url: str, titles: List[str]) -> Dict[str, dict]:
    """Get thumbnails in parallel."""
    results = {}
    
    with ThreadPoolExecutor(max_workers=getattr(config, "THUMBNAIL_WORKERS", 10)) as executor:
        future_to_title = {
            executor.submit(get_thumbnail_url, base_url, title): title
            for title in titles
        }
        
        for future in as_completed(future_to_title):
            title = future_to_title[future]
            try:
                thumb = future.result()
                if thumb:
                    results[title] = thumb
            except Exception:
                pass
    
    return results


def load_existing_images(filename: str) -> Set[str]:
    """Load already harvested image URLs from existing Excel file."""
    if not os.path.exists(filename):
        return set()
    
    try:
        wb = load_workbook(filename, read_only=True)
        ws = wb.active
        
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                existing.add(row[1])
        
        wb.close()
        return existing
    except PermissionError:
        print(f"[!] Cannot read '{filename}' - file may be open. Will start fresh.")
        return set()
    except Exception:
        return set()


def load_existing_titles(filename: str) -> Set[str]:
    """Load already harvested file titles from existing Excel file."""
    if not os.path.exists(filename):
        return set()
    
    try:
        wb = load_workbook(filename, read_only=True)
        ws = wb.active
        
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 7 and row[6]:
                existing.add(row[6])
        
        wb.close()
        return existing
    except PermissionError:
        print(f"[!] Cannot read '{filename}' - file may be open. Will start fresh.")
        return set()
    except Exception:
        return set()


def _parse_resolution_text(value: object) -> Optional[Tuple[int, int]]:
    if value is None:
        return None
    s = str(value)
    nums = re.findall(r"\d[\d,]*", s)
    if len(nums) < 2:
        return None
    try:
        w = int(nums[0].replace(",", ""))
        h = int(nums[1].replace(",", ""))
        return w, h
    except Exception:
        return None


def prune_existing_xlsx_by_resolution(
    filename: str,
    target_sizes: Optional[List[Tuple[int, int]]],
    tolerance: float,
    min_score: int,
) -> int:
    if not target_sizes or not os.path.exists(filename):
        return 0

    exact_only = bool(getattr(config, "EXACT_DIMENSIONS_ONLY", False))
    pixel_tolerance = int(getattr(config, "PIXEL_TOLERANCE", 0))
    if exact_only:
        pixel_tolerance = 0

    try:
        wb = load_workbook(filename)
        ws = wb.active
    except PermissionError:
        print(f"[!] Cannot write '{filename}' - skipping prune.")
        return 0
    except Exception:
        return 0

    removed = 0
    for row_idx in range(ws.max_row, 1, -1):
        parsed = _parse_resolution_text(ws.cell(row_idx, 6).value)
        if not parsed:
            continue

        w, h = parsed

        if exact_only or pixel_tolerance > 0:
            score = resolution_score_pixels(w, h, target_sizes, pixel_tolerance)
        else:
            score = resolution_score(w, h, target_sizes, tolerance)

        if score < min_score:
            ws.delete_rows(row_idx, 1)
            removed += 1

    if removed:
        try:
            wb.save(filename)
        except PermissionError:
            print(f"[!] Cannot save pruned '{filename}' - skipping.")

    try:
        wb.close()
    except Exception:
        pass

    return removed


def copy_to_viewing_file(source_file: str, viewing_file: str) -> bool:
    """Copy results file to viewing file for safe viewing during script execution."""
    if not os.path.exists(source_file):
        return False
    
    try:
        shutil.copy2(source_file, viewing_file)
        return True
    except PermissionError:
        # Viewing file might be open, try to write to a temp file and rename
        try:
            temp_file = viewing_file + ".tmp"
            shutil.copy2(source_file, temp_file)
            os.replace(temp_file, viewing_file)
            return True
        except Exception:
            # Just continue without copying - viewing file is in use
            return False
    except Exception:
        return False


def harvest_from_category(
    base_url: str,
    cat_title: str,
    max_items: int,
    target_sizes: Optional[List[Tuple[int, int]]],
    existing_urls: Set[str],
    existing_titles: Set[str],
    is_priority: bool = False,
    tolerance: float = 0.0,
    min_score: int = 90
) -> Tuple[List[dict], List[str]]:
    """Harvest from category, skipping already collected images. Returns (items, subcategories)."""
    items = []
    subcats = []
    
    priority_marker = "🏆 PRIORITY" if is_priority else "📁"
    print(f"\n[*] {priority_marker} Harvesting: {cat_title[:65]}")
    
    # Use generator for efficient category member fetching with continuation
    continue_token = None
    all_file_titles = []
    all_subcats = []
    
    while True:
        params = {
            "action": "query",
            "list": "categorymembers",
            "cmtitle": cat_title,
            "cmtype": "file|subcat",
            "cmlimit": "500",  # Max allowed
            "format": "json",
            "formatversion": "2",
        }
        
        if continue_token:
            params["cmcontinue"] = continue_token
        
        data = api_request(base_url, params)
        if not data:
            break
        
        members = data.get("query", {}).get("categorymembers", [])
        
        for m in members:
            if not isinstance(m, dict) or not m.get("title"):
                continue
                
            ns = m.get("ns")
            if ns == 6:  # File
                all_file_titles.append(m["title"])
            elif ns == 14:  # Category
                all_subcats.append(m["title"])
        
        # Check for continuation
        if "continue" in data and "cmcontinue" in data["continue"]:
            continue_token = data["continue"]["cmcontinue"]
        else:
            break
        
        # Limit for performance
        if len(all_file_titles) >= 1000:
            break
    
    subcats = all_subcats
    file_titles = all_file_titles
    
    if subcats:
        print(f"    📂 Found {len(subcats)} subcategories")
    
    if not file_titles:
        print(f"    ⊘ No files in this category")
        return [], subcats
    
    total_files = len(file_titles)
    print(f"    📄 Found {total_files} files")

    max_files_per_cat = int(getattr(config, "MAX_FILES_PER_CATEGORY", 0) or 0)
    large_sample = int(getattr(config, "LARGE_CATEGORY_SAMPLE", 0) or 0)
    if max_files_per_cat > 0 and total_files > max_files_per_cat and category_priority_key(cat_title) >= 2:
        if large_sample > 0:
            file_titles = file_titles[:large_sample]
        else:
            file_titles = file_titles[:50]

    scan_limit = int(getattr(config, "CATEGORY_FILE_SCAN_LIMIT", 0) or 0)
    if scan_limit > 0 and len(file_titles) > scan_limit:
        file_titles = file_titles[:scan_limit]
    
    exact_only = bool(getattr(config, "EXACT_DIMENSIONS_ONLY", False))
    pixel_tolerance = int(getattr(config, "PIXEL_TOLERANCE", 20))  # Default 20px
    if exact_only:
        pixel_tolerance = 0

    for i in range(0, len(file_titles), 50):
        if len(items) >= max_items:
            break
        
        batch = file_titles[i:i+50]
        sizes = fetch_batch_sizes(base_url, batch)
        
        selected_titles = []
        title_to_score = {}
        
        for title, sz in sizes.items():
            if title in existing_titles:
                continue
            
            if target_sizes:
                if exact_only or pixel_tolerance > 0:
                    score = resolution_score_pixels(sz["width"], sz["height"], target_sizes, pixel_tolerance)
                else:
                    score = resolution_score(sz["width"], sz["height"], target_sizes, tolerance)
                
                if score >= min_score:
                    selected_titles.append(title)
                    title_to_score[title] = score
            else:
                selected_titles.append(title)
                title_to_score[title] = 100
        
        if not selected_titles:
            continue
        
        batch_meta = fetch_batch_metadata(base_url, selected_titles)
        
        scored_items = []
        for title, meta in batch_meta.items():
            if len(items) >= max_items:
                break
            
            if meta["url"] in existing_urls:
                continue
            
            score = title_to_score.get(title, 0)
            if score < min_score:
                continue
            meta["resolution_score"] = score
            scored_items.append(meta)
        
        scored_items.sort(key=lambda x: x["resolution_score"], reverse=True)
        
        for meta in scored_items:
            if len(items) >= max_items:
                break
            items.append(meta)
            marker = "🏆" if is_priority else "✓"
            print(f"    {marker} [{len(items)}/{max_items}] {meta['width']}×{meta['height']} - {meta['title'][:40]}")
    
    if items:
        print(f"    ✅ Collected {len(items)} new images")
    
    return items, subcats


def append_to_xlsx(items: List[dict], filename: str, viewing_filename: str = None) -> None:
    """Append new images to existing Excel file or create new one, then copy to viewing file."""
    if not items:
        print("[!] No new items to save")
        return
    
    os.makedirs(os.path.dirname(os.path.abspath(filename)) or ".", exist_ok=True)
    
    file_exists = os.path.exists(filename)
    
    if file_exists:
        try:
            wb = load_workbook(filename)
            ws = wb.active
            print(f"[*] Appending {len(items)} images to: {filename}")
        except PermissionError:
            print(f"[!] Cannot write '{filename}' - trying backup...")
            backup_file = filename.replace(".xlsx", f"_backup_{int(time.time())}.xlsx")
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Images"
                headers = [
                    "image_page_url", "file_url", "line1", "line2",
                    "best_thumb_url", "best_res_under_1mb", "title",
                    "description", "license", "author"
                ]
                ws.append(headers)
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(1, col_idx).font = Font(bold=True)
                filename = backup_file
                file_exists = False
                print(f"[*] Creating backup file: {backup_file}")
            except Exception as e:
                print(f"[X] Failed to create backup: {e}")
                return
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Images"
        
        headers = [
            "image_page_url", "file_url", "line1", "line2",
            "best_thumb_url", "best_res_under_1mb", "title",
            "description", "license", "author"
        ]
        
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(1, col_idx).font = Font(bold=True)
        
        print(f"[*] Creating new file: {filename}")
    
    for it in items:
        author = it.get("author", "")
        license_type = it.get("license_type", "")
        desc = it.get("description", "")
        
        line1 = f"by {author}, {license_type}"
        line2 = desc[:64] if desc else ""
        
        best_res = ""
        if it.get("width") and it.get("height"):
            best_res = f"{it['width']:,} × {it['height']:,} pixels"
        
        row = [
            it.get("page_url", ""),
            it.get("url", ""),
            line1,
            line2,
            it.get("best_thumb_url", ""),
            best_res,
            it.get("title", ""),
            desc,
            license_type,
            author
        ]
        
        row_idx = ws.max_row + 1
        
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row_idx, col_idx, val)
            
            if col_idx in [1, 2, 5] and val:
                cell.hyperlink = val
                cell.font = Font(color="0000FF", underline="single")
    
    if not file_exists:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value or "")
                except:
                    val = ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 100)
    
    try:
        wb.save(filename)
        print(f"[+] Saved {len(items)} images to: {filename}")
    except PermissionError:
        print(f"[!] Cannot save '{filename}' - file is locked")
        return
    except Exception as e:
        print(f"[X] Failed to save: {e}")
        return
    
    # Copy to viewing file
    if viewing_filename:
        if copy_to_viewing_file(filename, viewing_filename):
            print(f"[+] Copied to viewing file: {viewing_filename}")
        else:
            print(f"[!] Could not update viewing file (may be open)")


def normalize_resolutions(config_value) -> Optional[List[Tuple[int, int]]]:
    """Normalize resolution config to list of tuples."""
    if config_value is None:
        return None
    
    if isinstance(config_value, list):
        result = []
        for item in config_value:
            if isinstance(item, (list, tuple)) and len(item) == 2:
                result.append((int(item[0]), int(item[1])))
        return result if result else None
    
    if isinstance(config_value, tuple) and len(config_value) == 2:
        return [(int(config_value[0]), int(config_value[1]))]
    
    return None


def main():
    parser = argparse.ArgumentParser(description="Wikimedia Commons Image Harvester for Set-Top Box")
    default_output = getattr(config, "DEFAULT_OUTPUT_FILE", "results.xlsx")
    viewing_output = getattr(config, "VIEWING_OUTPUT_FILE", "result_viewing.xlsx")
    parser.add_argument("--output", "-o", default=default_output, help=f"Output Excel file (default: {default_output})")
    parser.add_argument("--max", "-m", type=int, help="Override MAX_IMAGES from config")
    
    args = parser.parse_args()
    
    # Get viewing file path (same directory as output)
    output_dir = os.path.dirname(os.path.abspath(args.output))
    viewing_file = os.path.join(output_dir, viewing_output) if output_dir else viewing_output
    
    max_items = args.max if args.max else getattr(config, "MAX_IMAGES", 50)
    target_sizes = normalize_resolutions(getattr(config, "TARGET_RESOLUTION", None))
    runtime_tolerance = float(getattr(config, "TOLERANCE", 0.0))
    runtime_min_score = int(getattr(config, "MIN_RESOLUTION_SCORE", 95))
    batch_size = int(getattr(config, "BATCH_SIZE", 10))
    
    # Add common resolutions if allowed
    if target_sizes and getattr(config, "ALLOW_COMMON_RESOLUTIONS", False) and getattr(config, "COMMON_RESOLUTIONS", []):
        common = normalize_resolutions(getattr(config, "COMMON_RESOLUTIONS", []))
        if common:
            for res in common:
                if res not in target_sizes:
                    target_sizes.append(res)
    
    categories = getattr(config, "DEFAULT_CATEGORIES", [])
    if not categories:
        print("[X] No categories configured in config.py")
        sys.exit(1)
    
    print("=" * 70)
    print("🎨 WIKIMEDIA QUALITY IMAGE HARVESTER - SET-TOP BOX")
    print("   📍 Wiki Loves Monuments 2025 | Wiki Loves Folklore | Wiki Loves Birds")
    print("=" * 70)
    print(f"🎯 Target: {max_items} images")
    print(f"📁 Output: {args.output}")
    print(f"👁️  Viewing: {viewing_file}")
    print(f"📦 Batch size: {batch_size} images")
    
    exact_only = bool(getattr(config, "EXACT_DIMENSIONS_ONLY", False))
    pixel_tolerance = int(getattr(config, "PIXEL_TOLERANCE", 20))

    if target_sizes:
        if exact_only:
            print(f"📐 Resolutions (exact):")
        elif pixel_tolerance > 0:
            print(f"📐 Resolutions (±{pixel_tolerance}px):")
        else:
            print(f"📐 Resolutions (±{int(runtime_tolerance*100)}%):")
        for tw, th in target_sizes[:8]:
            print(f"   • {tw}×{th}")
        if len(target_sizes) > 8:
            print(f"   ... and {len(target_sizes) - 8} more")
    else:
        print(f"📐 Any resolution")
    
    removed_rows = prune_existing_xlsx_by_resolution(
        args.output,
        target_sizes,
        runtime_tolerance,
        runtime_min_score,
    )
    if removed_rows:
        print(f"🧹 Removed {removed_rows} existing rows outside target resolutions")

    existing_urls = load_existing_images(args.output)
    existing_titles = load_existing_titles(args.output)
    if existing_urls:
        print(f"📋 Found {len(existing_urls)} existing images (skipping duplicates)")
    if existing_titles:
        print(f"📋 Found {len(existing_titles)} existing titles (skipping duplicates)")
    
    print("=" * 70)
    
    # Harvest with PRIORITY
    all_items: List[dict] = []
    batch_buffer: List[dict] = []
    
    use_winning_priority = bool(getattr(config, "USE_WINNING_KEYWORD_PRIORITY", True))

    # Separate categories by priority
    priority_cats = []
    regular_cats = []
    
    for cat_url in categories:
        extracted = extract_category_title(cat_url)
        if not extracted:
            continue
        
        base_url, cat_title = extracted
        
        if use_winning_priority and is_winning_category(cat_title):
            priority_cats.append((base_url, cat_title, cat_url))
        else:
            regular_cats.append((base_url, cat_title, cat_url))
    
    # Initialize Queue: (base_url, cat_title, is_priority, depth)
    queue = deque()
    
    # Add priority categories first
    for base_url, cat_title, cat_url in priority_cats:
        if should_queue_category(cat_title):
            queue.append((base_url, cat_title, True, 0))
        
    # Add regular categories
    for base_url, cat_title, cat_url in regular_cats:
        if should_queue_category(cat_title):
            queue.append((base_url, cat_title, False, 0))
        
    visited_cats = set()
    print(f"\n🚀 Starting harvest with {len(queue)} categories in queue")
    
    max_depth = int(getattr(config, "MAX_SUBCATEGORY_DEPTH", 2))
    max_cats = int(getattr(config, "MAX_CATEGORIES_TO_SCAN", 0) or 0)
    max_without_match = int(getattr(config, "MAX_CATEGORIES_WITHOUT_MATCH", 0) or 0)
    scanned = 0
    without_match = 0

    while queue and len(all_items) < max_items:
        base_url, cat_title, is_prio, depth = queue.popleft()
        
        if cat_title in visited_cats:
            continue
        visited_cats.add(cat_title)

        scanned += 1
        if max_cats > 0 and scanned > max_cats:
            break
        if max_without_match > 0 and without_match >= max_without_match:
            break
        
        if depth > max_depth:
            continue
             
        remaining = max_items - len(all_items)
        
        items, subcats = harvest_from_category(
            base_url, cat_title, remaining, target_sizes, existing_urls, existing_titles, is_priority=is_prio,
            tolerance=runtime_tolerance, min_score=runtime_min_score
        )

        if not items:
            without_match += 1
        else:
            without_match = 0
        
        if items:
            all_items.extend(items)
            batch_buffer.extend(items)
            
            # Save batch if ready (every batch_size images)
            while len(batch_buffer) >= batch_size:
                to_save = batch_buffer[:batch_size]
                titles = [it["title"] for it in to_save]
                thumbs = process_thumbnails_batch(base_url, titles)
                for it in to_save:
                    thumb = thumbs.get(it["title"])
                    if thumb:
                        it["best_thumb_url"] = thumb.get("url")
                        it["best_thumb_width"] = thumb.get("width")
                        it["best_thumb_height"] = thumb.get("height")
                append_to_xlsx(to_save, args.output, viewing_file)
                for it in to_save:
                    if it.get("url"):
                        existing_urls.add(it["url"])
                    if it.get("title"):
                        existing_titles.add(it["title"])
                batch_buffer = batch_buffer[batch_size:]
        
        # Queue subcategories if we need more images
        if len(all_items) < max_items and depth < max_depth:
            next_depth = depth + 1
            filtered = [sc for sc in subcats if sc not in visited_cats and should_queue_category(sc)]
            filtered.sort(key=category_priority_key)
            max_subcats = int(getattr(config, "MAX_SUBCATEGORIES_PER_CATEGORY", 0) or 0)
            if max_subcats > 0:
                filtered = filtered[:max_subcats]
            for sc in filtered:
                queue.append((base_url, sc, is_prio, next_depth))
    
    if not all_items and not batch_buffer:
        print("\n[!] No new images found")
        if existing_urls:
            print(f"[*] Already have {len(existing_urls)} images in {args.output}")
        sys.exit(0)
    
    # Save any remaining items in buffer
    if batch_buffer:
        print(f"\n🖼️  Generating thumbnails for {len(batch_buffer)} remaining images...")
        base_url = "https://commons.wikimedia.org"
        titles = [it["title"] for it in batch_buffer]
        thumbs = process_thumbnails_batch(base_url, titles)
        for it in batch_buffer:
            thumb = thumbs.get(it["title"])
            if thumb:
                it["best_thumb_url"] = thumb.get("url")
                it["best_thumb_width"] = thumb.get("width")
                it["best_thumb_height"] = thumb.get("height")
        append_to_xlsx(batch_buffer, args.output, viewing_file)
        for it in batch_buffer:
            if it.get("url"):
                existing_urls.add(it["url"])
            if it.get("title"):
                existing_titles.add(it["title"])
    
    print(f"\n{'='*70}")
    print(f"✅ Added {len(all_items)} new images")
    
    print(f"📊 Total images in file: {len(existing_urls)}")
    print(f"📁 Results saved to: {args.output}")
    print(f"👁️  Viewing copy at: {viewing_file}")
    
    # Stats
    priority_count = sum(1 for it in all_items if it.get("resolution_score", 0) >= 80)
    print(f"🏆 High-quality matches: {priority_count}")
    
    res_counts = {}
    for it in all_items:
        res = f"{it['width']}×{it['height']}"
        res_counts[res] = res_counts.get(res, 0) + 1
    
    if res_counts:
        print(f"\n📐 New images by resolution:")
        for res, count in sorted(res_counts.items(), key=lambda x: -x[1])[:5]:
            print(f"   {res}: {count} images")
    
    print("=" * 70)


if __name__ == "__main__":
    main()
