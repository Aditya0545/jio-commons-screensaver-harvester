#!/usr/bin/env python3
"""
FAST MediaWiki Image Metadata Fetcher with FastAPI.

Optimizations over original script:
1. Async HTTP with aiohttp for parallel requests
2. Batch API queries (up to 50 titles per request)
3. Generator-based category member fetching with imageinfo in single call
4. Smart thumbnail selection without multiple HEAD requests
5. Connection pooling and session reuse
6. FastAPI REST endpoints for web integration

Requirements:
    pip install fastapi uvicorn aiohttp tqdm openpyxl pydantic

Usage (CLI mode):
    python get_image_metadata_fast.py
    python get_image_metadata_fast.py --csv results.csv --xlsx results.xlsx
    python get_image_metadata_fast.py "https://commons.wikimedia.org/wiki/Category:Wiki_Loves_Earth" --max 50

Usage (API server mode):
    python get_image_metadata_fast.py --serve
    # Then open http://localhost:8000/docs for interactive API docs
"""

import sys
import os
import csv
import argparse
import asyncio
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime, timezone
from urllib.parse import urlparse, unquote, parse_qs, quote
from dataclasses import dataclass, field, asdict
from enum import Enum

import aiohttp
from tqdm.asyncio import tqdm_asyncio
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:
    import config
    MEDIAWIKI_USERNAME = getattr(config, "MEDIAWIKI_USERNAME", "")
    DEFAULT_CATEGORIES = getattr(config, "DEFAULT_CATEGORIES", [])
    TARGET_RESOLUTION = getattr(config, "TARGET_RESOLUTION", None)
    MAX_IMAGES = getattr(config, "MAX_IMAGES", 0)
except ImportError:
    MEDIAWIKI_USERNAME = "WikiBot"
    DEFAULT_CATEGORIES = []
    TARGET_RESOLUTION = None
    MAX_IMAGES = 15

if not MEDIAWIKI_USERNAME or not str(MEDIAWIKI_USERNAME).strip():
    print("[X] MEDIAWIKI_USERNAME in config.py is empty. Please set it before running.")
    sys.exit(1)

# Wikimedia requires a detailed User-Agent with contact info
# Format: AppName/Version (Contact; URL) Library/Version
USER_AGENT = f"Wiki-Jio-Fast/2.0 (https://github.com/Aditya0545/jio-commons-screensaver-harvester; User:{MEDIAWIKI_USERNAME}) Python-aiohttp/3.9"

# Winner keywords for filtering - expanded list based on Wiki Loves competitions
WINNER_KEYWORDS = [
    # Top prizes
    "winner", "winners", "winning", "won",
    "first place", "first_place", "1st place", "1st_place",
    "second place", "second_place", "2nd place", "2nd_place",
    "third place", "third_place", "3rd place", "3rd_place",
    "fourth place", "fourth_place", "4th place", "5th place",
    "first prize", "first_prize", "1st prize",
    "second prize", "second_prize", "2nd prize",
    "third prize", "third_prize", "3rd prize",
    # Medal categories
    "gold", "silver", "bronze",
    "gold medal", "silver medal", "bronze medal",
    # Special mentions
    "honourable mention", "honorable mention",
    "special mention", "special prize", "special award",
    "jury prize", "jury selection", "jury's choice", "jury choice",
    "public choice", "people's choice", "popular vote", "audience choice",
    "editor's choice", "editors choice", "staff pick",
    # Finalist/shortlist
    "finalist", "finalists", "shortlist", "shortlisted",
    "top 10", "top 100", "top ten", "top hundred", "top 50", "top 20",
    "semi-finalist", "semifinalist", "quarterfinalist",
    # Featured/Quality on Wikimedia
    "featured picture", "featured_picture", "featured image",
    "quality image", "quality_image",
    "valued image", "valued_image",
    "picture of the day", "potd", "image of the day",
    "picture of the year", "poty", "image of the year",
    "picture of the month", "image of the month",
    "featured on commons", "fp on commons",
    # Wiki Loves specific campaigns
    "wiki loves earth", "wle",
    "wiki loves monuments", "wlm",
    "wiki loves africa", "wla",
    "wiki loves birds", "wlb",
    "wiki loves folklore",
    "wiki loves food",
    "wiki science competition", "wsc",
    "european science photo competition",
    # Country/region winners
    "national winner", "international winner",
    "country winner", "regional winner",
    "india winner", "indian winner",
    "usa winner", "american winner",
    "european winner", "asian winner", "african winner",
    # Year markers (often in winner categories)
    "2024", "2023", "2022", "2021", "2020", "2019", "2018", "2017",
    # Award types
    "award", "awarded", "prize", "prized", "best",
    "excellence", "outstanding", "remarkable",
    "selected", "chosen", "nominated", "nomination",
    # Nature/landscape specific
    "nature photograph", "landscape photograph",
    "wildlife photograph", "bird photograph",
    "monument photograph", "heritage photograph",
    # Quality indicators
    "high quality", "hq", "hi-res", "high resolution",
    "professional", "stunning", "beautiful", "magnificent",
    "breathtaking", "spectacular", "amazing",
]

# Keywords that indicate LOW quality or duplicates to SKIP
SKIP_KEYWORDS = [
    "crop", "cropped", "detail",
    "duplicate", "dupe",
    "thumbnail", "thumb",
    "low resolution", "low res", "lowres",
    "unfinished", "draft",
    "test", "example",
    "deleted", "nominated for deletion",
]

# Thumbnail widths to try (in priority order)
THUMB_WIDTHS = [2560, 1920, 1280, 1024, 800, 640, 512, 320, 256]
MAX_THUMB_SIZE = 1_000_000  # 1MB limit for thumbnails


@dataclass
class ImageMetadata:
    """Data class for image metadata."""
    batch: int = 0
    image_url: str = ""  # Page URL
    url: str = ""  # Highest resolution file URL
    line1: str = ""  # Author, License
    line2: str = ""  # Description (64 char limit)
    best_thumb_url: str = ""  # URL best resolution under 1MB
    best_res_under_1mb: str = ""  # Resolution string
    title: str = ""
    description: str = ""
    license: str = ""
    author: str = ""
    width: Optional[int] = None
    height: Optional[int] = None
    size_bytes: Optional[int] = None
    categories: List[str] = field(default_factory=list)
    creation_date: Optional[str] = None


class MediaWikiClient:
    """Async MediaWiki API client with connection pooling and batch queries."""
    
    def __init__(self, base_url: str, user_agent: str = USER_AGENT):
        self.base_url = base_url.rstrip('/')
        self.api_url = f"{self.base_url}/w/api.php"
        self.user_agent = user_agent
        self._session: Optional[aiohttp.ClientSession] = None
        self._connector: Optional[aiohttp.TCPConnector] = None
    
    async def __aenter__(self):
        await self._ensure_session()
        return self
    
    async def __aexit__(self, *args):
        await self.close()
    
    async def _ensure_session(self):
        if self._session is None or self._session.closed:
            self._connector = aiohttp.TCPConnector(
                limit=20,  # Max concurrent connections
                limit_per_host=10,
                ttl_dns_cache=300,
                enable_cleanup_closed=True,
            )
            timeout = aiohttp.ClientTimeout(total=30, connect=10)
            self._session = aiohttp.ClientSession(
                connector=self._connector,
                timeout=timeout,
                headers={
                    "User-Agent": self.user_agent,
                    "Api-User-Agent": self.user_agent,  # Wikimedia specifically checks this
                    "Accept": "application/json",
                },
            )
    
    async def close(self):
        if self._session and not self._session.closed:
            await self._session.close()
        if self._connector:
            await self._connector.close()
    
    async def api_request(self, params: Dict[str, Any], use_post: bool = False) -> Optional[Dict]:
        """Make an async API request with error handling.
        
        Args:
            params: Query parameters
            use_post: Use POST instead of GET (for large requests to avoid 414 URI Too Long)
        """
        await self._ensure_session()
        params.setdefault("format", "json")
        params.setdefault("formatversion", "2")
        
        try:
            if use_post:
                # POST request with form data to avoid URI length limits
                async with self._session.post(self.api_url, data=params) as resp:
                    if resp.status == 200:
                        return await resp.json()
                    else:
                        print(f"[X] API POST request failed: HTTP {resp.status}")
                        return None
            else:
                async with self._session.get(self.api_url, params=params) as resp:
                    if resp.status == 200:
                        return await resp.json()
                    elif resp.status == 414:
                        # URI too long, retry with POST
                        return await self.api_request(params, use_post=True)
                    else:
                        print(f"[X] API request failed: HTTP {resp.status}")
                        return None
        except asyncio.TimeoutError:
            print("[X] API request timed out")
            return None
        except Exception as e:
            print(f"[X] API request error: {e}")
            return None
    
    async def get_category_members_batch(
        self,
        category_title: str,
        cm_type: str = "file|subcat",
        limit: int = 500,
        continue_token: Optional[Dict] = None,
    ) -> Tuple[List[Dict], Optional[Dict]]:
        """
        Fetch category members in batch.
        Returns (members_list, continue_token_or_None).
        """
        params = {
            "action": "query",
            "list": "categorymembers",
            "cmtitle": category_title,
            "cmlimit": str(min(limit, 500)),
            "cmtype": cm_type,
            "cmprop": "ids|title|type",
        }
        if continue_token:
            params.update(continue_token)
        
        data = await self.api_request(params)
        if not data:
            return [], None
        
        members = data.get("query", {}).get("categorymembers", [])
        cont = data.get("continue")
        return members, cont
    
    async def get_imageinfo_batch(
        self,
        titles: List[str],
        thumb_width: int = 1280,
    ) -> Dict[str, Dict]:
        """
        Fetch imageinfo for multiple files at once.
        Uses POST requests to avoid URI length limits.
        Batch size is 20 to keep requests manageable.
        """
        if not titles:
            return {}
        
        # Use smaller batch size (20) to avoid URI too long errors
        # MediaWiki file titles can be very long
        BATCH_SIZE = 20
        results = {}
        
        for i in range(0, len(titles), BATCH_SIZE):
            batch = titles[i:i+BATCH_SIZE]
            titles_str = "|".join(batch)
            
            params = {
                "action": "query",
                "titles": titles_str,
                "prop": "imageinfo|categories",
                "iiprop": "timestamp|user|size|url|metadata|extmetadata|dimensions",
                "iiurlwidth": str(thumb_width),
                "cllimit": "50",
                "redirects": "true",
                "converttitles": "true",
            }
            
            # Use POST to avoid 414 URI Too Long
            data = await self.api_request(params, use_post=True)
            if not data:
                continue
            
            pages = data.get("query", {}).get("pages", [])
            for page in pages:
                if page.get("missing"):
                    continue
                title = page.get("title", "")
                results[title] = page
        
        return results
    
    async def get_category_files_with_imageinfo(
        self,
        category_title: str,
        max_items: Optional[int] = None,
        target_size: Optional[Tuple[int, int]] = None,
        filter_winners: bool = True,
        thumb_width: int = 1280,
        max_categories: int = 100,  # Limit to prevent infinite traversal
        max_depth: int = 5,  # Max subcategory depth
    ) -> List[ImageMetadata]:
        """
        Recursively fetch category files with their imageinfo using generators.
        This combines category traversal with metadata fetching for efficiency.
        Includes deduplication and quality-based sorting.
        """
        seen_cats: set = set()
        seen_files: set = set()
        seen_signatures: set = set()  # For deduplication
        items: List[ImageMetadata] = []
        batch_counter = [0]  # Use list for mutable closure
        
        # Track duplicates skipped
        duplicates_skipped = [0]
        
        # Progress tracking
        total_files_scanned = [0]
        total_files_checked = [0]
        total_winners_found = [0]
        
        def print_progress(extra: str = ""):
            """Print current progress on same line."""
            target_str = f" (target: {max_items})" if max_items else ""
            size_str = f" | Res: {target_size[0]}x{target_size[1]}" if target_size else ""
            filter_str = " | 🏆 winners only" if filter_winners else " | all images"
            dup_str = f" | 🔄 {duplicates_skipped[0]} dupes" if duplicates_skipped[0] > 0 else ""
            msg = f"\r  📊 {len(seen_cats)}/{max_categories} cats | {total_files_scanned[0]} files | {total_files_checked[0]} checked | ✅ {len(items)} matched{dup_str}{target_str}{size_str}{filter_str} {extra}"
            print(msg, end="", flush=True)
        
        def add_item_if_unique(meta: ImageMetadata) -> bool:
            """Add item only if it's not a duplicate. Returns True if added."""
            signature = get_image_signature(meta)
            
            if signature in seen_signatures:
                duplicates_skipped[0] += 1
                return False
            
            # Check if we have a similar image - keep the one with higher quality score
            seen_signatures.add(signature)
            
            # Calculate quality score
            meta.batch = get_image_quality_score(meta)  # Repurpose batch as quality score temporarily
            
            items.append(meta)
            return True
        
        # Collect all file titles first, then batch fetch metadata
        file_queue: List[str] = []
        # Queue contains (category_title, depth) tuples
        subcat_queue: List[Tuple[str, int]] = [(category_title, 0)]
        
        print(f"[*] Starting category traversal: {category_title}")
        print(f"    📋 Limits: max {max_items} images, max {max_categories} categories, max depth {max_depth}")
        if target_size:
            print(f"    🔍 Resolution filter: {target_size[0]}x{target_size[1]}")
        if filter_winners:
            print(f"    🏆 Winner filter: ON (only images with 'winner', 'gold', 'prize' etc.)")
        else:
            print(f"    📷 Winner filter: OFF (fetching all images)")
        print()
        
        while subcat_queue and (max_items is None or len(items) < max_items):
            # Check category limit
            if len(seen_cats) >= max_categories:
                print(f"\n\n[!] Reached max categories limit ({max_categories}). Stopping traversal.")
                break
            
            current_cat, depth = subcat_queue.pop(0)
            if current_cat in seen_cats:
                continue
            if depth > max_depth:
                continue  # Skip categories that are too deep
            
            seen_cats.add(current_cat)
            
            # Show which category we're scanning
            cat_short = current_cat.replace("Category:", "")[:50]
            print(f"\n  📁 [{depth}] {cat_short}{'...' if len(current_cat) > 60 else ''}")
            
            # Fetch all members of this category
            cont = None
            page_num = 0
            while True:
                page_num += 1
                members, cont = await self.get_category_members_batch(
                    current_cat, limit=500, continue_token=cont
                )
                
                new_files = 0
                new_subcats = 0
                for m in members:
                    mtype = m.get("type")
                    mtitle = m.get("title", "")
                    
                    if mtype == "subcat" and mtitle not in seen_cats:
                        subcat_queue.append((mtitle, depth + 1))
                        new_subcats += 1
                    elif mtype == "file" and mtitle not in seen_files:
                        seen_files.add(mtitle)
                        file_queue.append(mtitle)
                        total_files_scanned[0] += 1
                        new_files += 1
                
                if new_files > 0 or new_subcats > 0:
                    print_progress(f"| +{new_files} files, +{new_subcats} subcats")
                
                # Process files in batches of 20 to get early results
                while len(file_queue) >= 20 or (not cont and file_queue):
                    batch_titles = file_queue[:20]
                    file_queue = file_queue[20:]
                    
                    print_progress(f"| Fetching metadata...")
                    page_data = await self.get_imageinfo_batch(batch_titles, thumb_width)
                    
                    for title, page in page_data.items():
                        meta = self._parse_imageinfo(page, batch_counter[0])
                        batch_counter[0] += 1
                        total_files_checked[0] += 1
                        
                        if meta is None:
                            continue
                        
                        # Apply filters
                        if filter_winners and not is_winner_metadata(meta):
                            continue
                        
                        if target_size:
                            tw, th = target_size
                            if meta.width != tw or meta.height != th:
                                continue
                        
                        # Add with deduplication check
                        if add_item_if_unique(meta):
                            score = get_image_quality_score(meta)
                            print(f"\n    ✅ #{len(items)} [score:{score}]: {meta.title[:55]}...")
                        
                        if max_items and len(items) >= max_items:
                            print(f"\n\n[+] ✅ Reached max items: {max_items}")
                            # Sort by quality score before returning
                            items.sort(key=lambda x: x.batch, reverse=True)
                            # Reset batch numbers
                            for i, item in enumerate(items):
                                item.batch = i + 1
                            return items
                    
                    print_progress()
                    
                    if not cont and not file_queue:
                        break
                
                if not cont:
                    break
            
            # Show queue status periodically
            if len(subcat_queue) > 0:
                print_progress(f"| {len(subcat_queue)} subcats remaining")
        
        # Process remaining files
        if file_queue:
            print(f"\n  📦 Processing {len(file_queue)} remaining files...")
        
        while file_queue and (max_items is None or len(items) < max_items):
            batch_titles = file_queue[:20]
            file_queue = file_queue[20:]
            
            print_progress(f"| Fetching final batches...")
            page_data = await self.get_imageinfo_batch(batch_titles, thumb_width)
            
            for title, page in page_data.items():
                meta = self._parse_imageinfo(page, batch_counter[0])
                batch_counter[0] += 1
                total_files_checked[0] += 1
                
                if meta is None:
                    continue
                
                if filter_winners and not is_winner_metadata(meta):
                    continue
                
                if target_size:
                    tw, th = target_size
                    if meta.width != tw or meta.height != th:
                        continue
                
                # Add with deduplication check
                if add_item_if_unique(meta):
                    score = get_image_quality_score(meta)
                    print(f"\n    ✅ #{len(items)} [score:{score}]: {meta.title[:55]}...")
                
                if max_items and len(items) >= max_items:
                    break
            
            print_progress()
        
        # Sort by quality score (stored in batch field) and reset batch numbers
        items.sort(key=lambda x: x.batch, reverse=True)
        for i, item in enumerate(items):
            item.batch = i + 1
        
        print(f"\n\n[*] Traversal complete: {len(seen_cats)} categories, {total_files_scanned[0]} files scanned")
        print(f"    ✅ {len(items)} unique images found, {duplicates_skipped[0]} duplicates skipped")
        return items
    
    def _parse_imageinfo(self, page: Dict, batch_num: int = 0) -> Optional[ImageMetadata]:
        """Parse API response page into ImageMetadata."""
        imageinfo = page.get("imageinfo", [])
        if not imageinfo:
            return None
        
        info = imageinfo[0]
        ext = info.get("extmetadata", {}) or {}
        
        def ext_value(key: str) -> Optional[str]:
            val = ext.get(key)
            if isinstance(val, dict):
                return val.get("value")
            return val
        
        title = page.get("title", "")
        page_url = f"{self.base_url}/wiki/{quote(title.replace(' ', '_'), safe=':/')}"
        file_url = info.get("url", "")
        
        author = ext_value("Artist") or info.get("user", "")
        license_type = ext_value("LicenseShortName") or ext_value("License") or ""
        description = ext_value("ImageDescription") or ext_value("ObjectName") or ""
        creation_date = ext_value("DateTimeOriginal") or ext_value("DateTime")
        
        width = info.get("width")
        height = info.get("height")
        size_bytes = info.get("size")
        
        categories_raw = page.get("categories", []) or []
        categories = [c.get("title", "") for c in categories_raw if isinstance(c, dict)]
        
        # Get thumbnail info from response
        thumb_url = info.get("thumburl", "")
        thumb_width = info.get("thumbwidth")
        thumb_height = info.get("thumbheight")
        
        # Format line1 and line2
        line1 = f"by {author}, {license_type}" if author or license_type else ""
        line2 = (description[:64]).replace("\n", " ").replace("\r", "") if description else ""
        
        best_res = ""
        if thumb_width and thumb_height:
            best_res = f"{thumb_width} × {thumb_height} pixels"
        
        return ImageMetadata(
            batch=batch_num,
            image_url=page_url,
            url=file_url,
            line1=line1,
            line2=line2,
            best_thumb_url=thumb_url,
            best_res_under_1mb=best_res,
            title=title,
            description=description,
            license=license_type,
            author=author,
            width=width,
            height=height,
            size_bytes=size_bytes,
            categories=categories,
            creation_date=creation_date,
        )
    
    async def find_best_thumbnail(
        self,
        title: str,
        max_size: int = MAX_THUMB_SIZE,
    ) -> Optional[Dict[str, Any]]:
        """
        Find the best thumbnail under max_size bytes.
        Uses binary search approach with parallel requests.
        """
        # Query multiple widths in parallel to find the best one
        tasks = []
        for width in THUMB_WIDTHS[:5]:  # Check top 5 widths first
            params = {
                "action": "query",
                "titles": title,
                "prop": "imageinfo",
                "iiprop": "url|size",
                "iiurlwidth": str(width),
            }
            tasks.append(self.api_request(params))
        
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        for i, result in enumerate(results):
            if isinstance(result, Exception) or not result:
                continue
            
            pages = result.get("query", {}).get("pages", [])
            if not pages:
                continue
            
            page = pages[0]
            imageinfo = page.get("imageinfo", [])
            if not imageinfo:
                continue
            
            info = imageinfo[0]
            thumb_url = info.get("thumburl") or info.get("url")
            thumb_w = info.get("thumbwidth")
            thumb_h = info.get("thumbheight")
            
            if thumb_url:
                # Estimate size based on resolution (rough heuristic)
                # For accurate size, we'd need HEAD request, but this is faster
                estimated_size = (thumb_w or 0) * (thumb_h or 0) * 0.5  # ~0.5 bytes per pixel for JPEG
                if estimated_size < max_size:
                    return {
                        "url": thumb_url,
                        "width": thumb_w,
                        "height": thumb_h,
                        "estimated_size": estimated_size,
                    }
        
        return None


def is_winner_metadata(meta: ImageMetadata) -> bool:
    """Check if metadata indicates a winning image."""
    title_lower = meta.title.lower()
    
    # First check if it should be SKIPPED (low quality indicators)
    for kw in SKIP_KEYWORDS:
        if kw in title_lower:
            return False
    
    # Check categories for skip keywords
    for cat in meta.categories:
        cat_lower = cat.lower()
        for kw in SKIP_KEYWORDS:
            if kw in cat_lower:
                return False
    
    # Now check for winner keywords in categories
    for cat in meta.categories:
        cat_lower = cat.lower()
        for kw in WINNER_KEYWORDS:
            if kw in cat_lower:
                return True
    
    # Check title for winner keywords
    for kw in WINNER_KEYWORDS:
        if kw in title_lower:
            return True
    
    return False


def get_image_quality_score(meta: ImageMetadata) -> int:
    """
    Calculate a quality score for an image to help prioritize better images.
    Higher score = better quality.
    """
    score = 0
    title_lower = meta.title.lower()
    cats_lower = " ".join(meta.categories).lower()
    combined = title_lower + " " + cats_lower
    
    # Resolution score (prefer higher resolution)
    if meta.width and meta.height:
        pixels = meta.width * meta.height
        if pixels >= 24000000:  # 24MP+
            score += 50
        elif pixels >= 12000000:  # 12MP+
            score += 40
        elif pixels >= 6000000:  # 6MP+
            score += 30
        elif pixels >= 2000000:  # 2MP+
            score += 20
        else:
            score += 10
    
    # Prize level score
    if any(kw in combined for kw in ["first place", "1st place", "first prize", "1st prize", "gold"]):
        score += 100
    elif any(kw in combined for kw in ["second place", "2nd place", "second prize", "2nd prize", "silver"]):
        score += 80
    elif any(kw in combined for kw in ["third place", "3rd place", "third prize", "3rd prize", "bronze"]):
        score += 60
    elif any(kw in combined for kw in ["finalist", "shortlist", "top 10", "top 100"]):
        score += 40
    elif any(kw in combined for kw in ["winner", "winning"]):
        score += 50
    elif any(kw in combined for kw in ["honourable mention", "honorable mention", "special mention"]):
        score += 30
    
    # Featured/Quality badges
    if "featured picture" in combined or "featured_picture" in combined:
        score += 70
    if "quality image" in combined or "quality_image" in combined:
        score += 50
    if "valued image" in combined or "valued_image" in combined:
        score += 40
    
    # Has good description
    if meta.description and len(meta.description) > 50:
        score += 20
    elif meta.description and len(meta.description) > 20:
        score += 10
    
    # Has author info
    if meta.author and len(meta.author) > 2:
        score += 10
    
    # Has license info
    if meta.license:
        score += 5
    
    return score


def get_image_signature(meta: ImageMetadata) -> str:
    """
    Generate a signature to detect similar/duplicate images.
    Images with same signature are considered duplicates.
    """
    # Use author + approximate resolution + first part of description
    author = (meta.author or "unknown").lower().strip()[:30]
    
    # Round resolution to nearest 1000 to group similar sizes
    w = (meta.width or 0) // 1000 * 1000
    h = (meta.height or 0) // 1000 * 1000
    
    # First 50 chars of description (normalized)
    desc = (meta.description or "").lower().replace("\n", " ").strip()[:50]
    
    # Remove common prefixes from title to find base name
    title = meta.title.lower()
    for prefix in ["file:", "image:"]:
        if title.startswith(prefix):
            title = title[len(prefix):]
    
    # Extract base filename (without numbers/suffixes)
    import re
    base_title = re.sub(r'[\d_\-\.\s]+\.(jpg|jpeg|png|gif|tiff|svg)$', '', title, flags=re.IGNORECASE)
    base_title = re.sub(r'\s*\(\d+\)\s*', '', base_title)  # Remove (1), (2), etc.
    base_title = base_title[:40]
    
    return f"{author}|{w}x{h}|{base_title}"


def extract_category_from_url(url: str) -> Optional[Tuple[str, str]]:
    """Extract base_url and category title from a category URL."""
    parsed = urlparse(url)
    base_url = f"{parsed.scheme}://{parsed.netloc}"
    
    if "/wiki/" in parsed.path:
        cat_title = unquote(parsed.path.split("/wiki/", 1)[1])
    elif "title=" in parsed.query:
        query_params = parse_qs(parsed.query)
        cat_title = unquote(query_params.get("title", [""])[0])
    else:
        return None
    
    if not cat_title.startswith("Category:"):
        return None
    
    return base_url, cat_title


# ========== OUTPUT HELPERS ==========

def save_to_csv(items: List[ImageMetadata], filename: str) -> None:
    """Save metadata to CSV file."""
    if not items:
        print("[!] No items to save to CSV.")
        return
    
    os.makedirs(os.path.dirname(os.path.abspath(filename)) or ".", exist_ok=True)
    
    headers = [
        "batch", "image_url", "url", "line1", "line2",
        "best_thumb_url", "best_res_under_1mb", "title",
        "description", "license", "author"
    ]
    
    try:
        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for item in items:
                writer.writerow([
                    item.batch,
                    item.image_url,
                    item.url,
                    item.line1,
                    item.line2,
                    item.best_thumb_url,
                    item.best_res_under_1mb,
                    item.title,
                    item.description,
                    item.license,
                    item.author,
                ])
        print(f"[+] Saved {len(items)} rows to CSV: {filename}")
    except Exception as e:
        print(f"[X] Failed to write CSV: {e}")


def save_to_xlsx(items: List[ImageMetadata], filename: str) -> None:
    """Save metadata to XLSX with hyperlinks."""
    if not items:
        print("[!] No items to save to XLSX.")
        return
    
    os.makedirs(os.path.dirname(os.path.abspath(filename)) or ".", exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "images"
    
    headers = [
        "batch", "image_url", "url", "line1", "line2",
        "best_thumb_url", "best_res_under_1mb", "title",
        "description", "license", "author"
    ]
    
    header_font = Font(bold=True)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = header_font
    
    link_font = Font(color="0000FF", underline="single")
    
    for item in items:
        row_idx = ws.max_row + 1
        
        ws.cell(row=row_idx, column=1, value=item.batch)
        
        # image_url with hyperlink
        cell = ws.cell(row=row_idx, column=2, value=item.image_url)
        if item.image_url:
            cell.hyperlink = item.image_url
            cell.font = link_font
        
        # url with hyperlink
        cell = ws.cell(row=row_idx, column=3, value=item.url)
        if item.url:
            cell.hyperlink = item.url
            cell.font = link_font
        
        ws.cell(row=row_idx, column=4, value=item.line1)
        ws.cell(row=row_idx, column=5, value=item.line2)
        
        # best_thumb_url with hyperlink
        cell = ws.cell(row=row_idx, column=6, value=item.best_thumb_url)
        if item.best_thumb_url:
            cell.hyperlink = item.best_thumb_url
            cell.font = link_font
        
        ws.cell(row=row_idx, column=7, value=item.best_res_under_1mb)
        ws.cell(row=row_idx, column=8, value=item.title)
        ws.cell(row=row_idx, column=9, value=item.description)
        ws.cell(row=row_idx, column=10, value=item.license)
        ws.cell(row=row_idx, column=11, value=item.author)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
    
    try:
        wb.properties.creator = USER_AGENT
        wb.properties.created = datetime.now(timezone.utc)
    except Exception:
        pass
    
    try:
        wb.save(filename)
        print(f"[+] Saved {len(items)} rows to XLSX: {filename}")
    except PermissionError:
        print(f"[X] Permission denied. Is '{filename}' open in Excel?")
    except Exception as e:
        print(f"[X] Failed to write XLSX: {e}")


# ========== FASTAPI SERVER ==========

def create_fastapi_app():
    """Create and configure FastAPI application."""
    from fastapi import FastAPI, Query, HTTPException, BackgroundTasks
    from fastapi.responses import FileResponse, JSONResponse
    from pydantic import BaseModel, Field
    
    app = FastAPI(
        title="MediaWiki Image Metadata API",
        description="Fast async API for fetching MediaWiki image metadata",
        version="2.0.0",
    )
    
    class CategoryRequest(BaseModel):
        url: str = Field(..., description="Category URL to fetch images from")
        max_items: Optional[int] = Field(15, description="Maximum number of images to fetch")
        filter_winners: bool = Field(True, description="Only return contest winners")
        target_width: Optional[int] = Field(None, description="Filter by image width")
        target_height: Optional[int] = Field(None, description="Filter by image height")
    
    class ImageResponse(BaseModel):
        batch: int
        image_url: str
        url: str
        line1: str
        line2: str
        best_thumb_url: str
        best_res_under_1mb: str
        title: str
        description: str
        license: str
        author: str
    
    @app.get("/")
    async def root():
        return {"message": "MediaWiki Image Metadata API", "docs": "/docs"}
    
    @app.get("/health")
    async def health():
        return {"status": "healthy"}
    
    @app.post("/fetch-category", response_model=List[ImageResponse])
    async def fetch_category(request: CategoryRequest):
        """Fetch images from a MediaWiki category."""
        extracted = extract_category_from_url(request.url)
        if not extracted:
            raise HTTPException(400, "Invalid category URL")
        
        base_url, category_title = extracted
        target_size = None
        if request.target_width and request.target_height:
            target_size = (request.target_width, request.target_height)
        
        async with MediaWikiClient(base_url) as client:
            items = await client.get_category_files_with_imageinfo(
                category_title,
                max_items=request.max_items,
                target_size=target_size,
                filter_winners=request.filter_winners,
            )
        
        return [
            ImageResponse(
                batch=item.batch,
                image_url=item.image_url,
                url=item.url,
                line1=item.line1,
                line2=item.line2,
                best_thumb_url=item.best_thumb_url,
                best_res_under_1mb=item.best_res_under_1mb,
                title=item.title,
                description=item.description,
                license=item.license,
                author=item.author,
            )
            for item in items
        ]
    
    @app.get("/fetch-category")
    async def fetch_category_get(
        url: str = Query(..., description="Category URL"),
        max_items: int = Query(15, description="Max items"),
        filter_winners: bool = Query(True, description="Filter winners only"),
    ):
        """GET endpoint for fetching category images."""
        extracted = extract_category_from_url(url)
        if not extracted:
            raise HTTPException(400, "Invalid category URL")
        
        base_url, category_title = extracted
        
        async with MediaWikiClient(base_url) as client:
            items = await client.get_category_files_with_imageinfo(
                category_title,
                max_items=max_items,
                filter_winners=filter_winners,
            )
        
        return [asdict(item) for item in items]
    
    return app


# ========== CLI ==========

async def main_async(args) -> None:
    """Async main function for CLI mode."""
    target_size = None
    if args.width and args.height:
        target_size = (args.width, args.height)
    elif TARGET_RESOLUTION and isinstance(TARGET_RESOLUTION, (list, tuple)) and len(TARGET_RESOLUTION) == 2:
        target_size = (int(TARGET_RESOLUTION[0]), int(TARGET_RESOLUTION[1]))
    
    effective_max = args.max if args.max != 15 else (MAX_IMAGES if MAX_IMAGES > 0 else None)
    
    # Get filter settings from args
    filter_winners = not getattr(args, 'no_filter', False)
    max_cats = getattr(args, 'max_cats', 50)
    max_depth = getattr(args, 'max_depth', 3)
    
    items: List[ImageMetadata] = []
    
    if args.url:
        extracted = extract_category_from_url(args.url)
        if not extracted:
            print("[X] Invalid category URL")
            sys.exit(1)
        
        base_url, category_title = extracted
        
        async with MediaWikiClient(base_url) as client:
            items = await client.get_category_files_with_imageinfo(
                category_title,
                max_items=effective_max,
                target_size=target_size,
                filter_winners=filter_winners,
                max_categories=max_cats,
                max_depth=max_depth,
            )
    else:
        if not DEFAULT_CATEGORIES:
            print("[X] No URL provided and DEFAULT_CATEGORIES is empty.")
            sys.exit(1)
        
        remaining = effective_max
        for cat_url in DEFAULT_CATEGORIES:
            if remaining is not None and remaining <= 0:
                break
            
            extracted = extract_category_from_url(cat_url)
            if not extracted:
                continue
            
            base_url, category_title = extracted
            cat_max = remaining
            
            async with MediaWikiClient(base_url) as client:
                found = await client.get_category_files_with_imageinfo(
                    category_title,
                    max_items=cat_max,
                    target_size=target_size,
                    filter_winners=filter_winners,
                    max_categories=max_cats,
                    max_depth=max_depth,
                )
            
            for item in found:
                items.append(item)
                if remaining is not None:
                    remaining -= 1
                    if remaining <= 0:
                        break
    
    if not items:
        print("[X] No images found matching criteria.")
        print("[*] TIP: Try --no-filter to get all images, or adjust --max-cats and --max-depth")
        sys.exit(1)
    
    print(f"\n[+] Found {len(items)} images")
    for item in items:
        print(f"  - {item.title}")
    
    if args.csv:
        save_to_csv(items, args.csv)
    
    xlsx_path = args.xlsx or os.path.join(os.getcwd(), "results.xlsx")
    save_to_xlsx(items, xlsx_path)


def main() -> None:
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Fast MediaWiki Image Metadata Fetcher with FastAPI"
    )
    parser.add_argument("url", nargs="?", help="Category URL")
    parser.add_argument("--csv", "-c", metavar="FILE", help="Output CSV file")
    parser.add_argument("--xlsx", "-x", metavar="FILE", help="Output XLSX file")
    parser.add_argument("--max", "-m", type=int, default=15, help="Max items")
    parser.add_argument("--width", type=int, help="Target width filter")
    parser.add_argument("--height", type=int, help="Target height filter")
    parser.add_argument("--no-filter", "-n", action="store_true", help="Disable winner filtering (get ALL images)")
    parser.add_argument("--max-cats", type=int, default=50, help="Max categories to scan (default: 50)")
    parser.add_argument("--max-depth", type=int, default=3, help="Max subcategory depth (default: 3)")
    parser.add_argument("--serve", "-s", action="store_true", help="Run as FastAPI server")
    parser.add_argument("--port", "-p", type=int, default=8000, help="Server port")
    
    args = parser.parse_args()
    
    if args.serve:
        import uvicorn
        app = create_fastapi_app()
        print(f"[*] Starting FastAPI server on http://localhost:{args.port}")
        print(f"[*] API docs available at http://localhost:{args.port}/docs")
        uvicorn.run(app, host="0.0.0.0", port=args.port)
    else:
        asyncio.run(main_async(args))


if __name__ == "__main__":
    main()
